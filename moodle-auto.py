import os
import openpyxl
import re

from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
from datetime import datetime

# from selenium.webdriver.common.keys import Keys
from dotenv import load_dotenv

load_dotenv()  # .envファイルから環境変数を読み込む

student_id = os.getenv("STUDENT_ID")
password = os.getenv("PASSWORD")


# moodleにアクセスし、学生番号とパスワードを入力してログインする関数
def login_to_moodle(driver, student_id, password):
    driver.get("https://moodle.vle.hiroshima-u.ac.jp/local/hulogin/index.php")
    # 画面の最大化(ウィンドサイズに依存する問題を避ける)
    driver.maximize_window()
    sleep(1)

    driver.find_element(By.XPATH, '//a[contains(., "広大IDでログイン")]').click()
    input_id = driver.find_element(By.XPATH, '//input[@id="username"]')
    input_id.send_keys(student_id)
    input_password = driver.find_element(By.XPATH, '//input[@id="password"]')
    input_password.send_keys(password)
    login_button = driver.find_element(By.XPATH, '//button[@name="_eventId_proceed"]')
    login_button.click()
    sleep(5)

    print("ログイン完了")


# 指定した科目の指定したスレッドにアクセスする関数
def access_assignment(driver, course_name, section_title, title):
    driver.find_element(By.XPATH, "//a[contains(normalize-space(), 'Home')]").click()
    driver.find_element(By.XPATH, f'//a[contains(., "{course_name}")]').click()
    driver.find_element(By.XPATH, f'//a[contains(., "{section_title}")]').click()
    sleep(1)
    driver.find_element(
        By.XPATH, f"//a[contains(normalize-space(), '{title}')]"
    ).click()

    print("課題にアクセス完了")


# 新規excelファイルの作成とエラー処理を行う関数
def create_or_load_excel(filepath):
    if not os.path.exists(filepath):
        print("ファイルがありませんので作成します")
        wb = openpyxl.Workbook()
        print("excelファイルの作成が完了しました")
        wb.save(filepath)
    else:
        wb = openpyxl.load_workbook(filepath)
        print("既に作成されています")
    return wb


# 各スレッドにアクセスして、class_nameを含む要素の数を返す関数
# TODO: クラスやTagは異なる可能性があるので開発者ツールで確認し、適宜変更
def threads_list(driver, class_name):
    thread_entry_button = driver.find_element(By.CLASS_NAME, "indent").find_elements(
        By.TAG_NAME, "article"  # スレッド
    )
    elements_with_class = [
        element
        for element in thread_entry_button
        if class_name in element.get_attribute("class")
    ]
    return elements_with_class


# 苗字と名前が反転している場合の対応
def names_match(name1, name2):
    name1_words = set(name1.split())
    name2_words = set(name2.split())
    return name1_words == name2_words


# 提出しているかどうかをexcelに書き込む関数
def submit_info_write_excel(
    search_word,
    student_name,
    after_check_filepath,
    after_check_file_sheet,
    submit_time,
    submit_deadline,
    delay_deadline,
):
    wb = openpyxl.load_workbook(after_check_filepath)

    if after_check_file_sheet not in wb.sheetnames:
        print("シートが存在しない")
        return

    sheet = wb[after_check_file_sheet]

    # 課題名に一致する列を探す
    # TODO: 初回に作成したファイルの課題のスレッド開始位置によって変更する
    for i in range(9, 25):  # 課題（横）のループ
        cell_word = sheet.cell(row=2, column=i).value

        if (
            (str(cell_word) in str(search_word))
            or (str(search_word) in str(cell_word))
            or (str(cell_word) == str(search_word))
        ):
            # 学生の名前に一致する行を探す
            # TODO: 名簿の学生の数によって変更する
            # 3行目から97行目までのループ
            for j in range(3, 98):  # 学生（縦）のループ
                # columnは学生の名前の列
                cell_name = sheet.cell(row=j, column=4).value
                if names_match(student_name, cell_name):
                    cell_name = sheet.cell(row=j, column=4).value
                    if names_match(student_name, cell_name):
                        submitted = sheet.cell(row=j, column=i).value
                        deadline = datetime.strptime(
                            submit_deadline, "%Y-%m-%d %H:%M:%S"
                        )
                        delay = datetime.strptime(delay_deadline, "%Y-%m-%d %H:%M:%S")

                        if submit_time < deadline and submitted != 1:
                            sheet.cell(row=j, column=i).value = 1
                            print(f"課題 '{search_word}' に {student_name} が提出しました。")
                        elif deadline < submit_time < delay and submitted != 1:
                            sheet.cell(row=j, column=i).value = 0.8
                            print(f"課題 '{search_word}' に {student_name} が遅延提出しました。")
                        break
            break
        else:
            print(f"課題 '{search_word}' は'{cell_word}'とは一致しません")

    wb.save(after_check_filepath)


# 各スレッドの情報を取得して新規excelに書き込む関数
def process_thread(
    thread_element,
    workbook,
    thread_list_filepath,
    after_check_filepath,
    after_check_file_sheet,
    submit_deadline,
    delay_deadline,
):
    thread_title = (
        thread_element.find_element(By.TAG_NAME, "header")
        .find_element(By.TAG_NAME, "h3")
        .text.translate(str.maketrans({"?": "", ":": "", "？": "", "　": "", " ": ""}))[
            :10
        ]
    )
    thread_title_original = (
        thread_element.find_element(By.TAG_NAME, "header")
        .find_element(By.TAG_NAME, "h3")
        .text
    )
    name = (
        thread_element.find_element(By.TAG_NAME, "header")
        .find_element(By.TAG_NAME, "a")
        .text
    )
    temp_time = (
        thread_element.find_element(By.TAG_NAME, "header")
        .find_element(By.TAG_NAME, "time")
        .text
    )
    # datatimeで格納するために、正規表現で整形
    match = re.search(r"(\d{4})年 (\d{1,2})月 (\d{1,2})日.* (\d{1,2}:\d{2})", temp_time)
    if match:
        time = datetime.strptime(
            f"{match.group(1)}-{match.group(2)}-{match.group(3)} {match.group(4)}",
            "%Y-%m-%d %H:%M",
        )
    else:
        print("日付の取得に失敗しました")

    user_info = [name, thread_title, time]

    print(user_info)

    # シートの確認と作成
    if thread_title in workbook.sheetnames:
        ws = workbook[thread_title]
    else:
        ws = workbook.create_sheet(index=0, title=thread_title)

    # データの書き込み
    row = ws.max_row + 1
    ws.cell(row=row, column=1).value = user_info[0]  # 名前
    ws.cell(row=row, column=2).value = user_info[1]  # タイトル
    ws.cell(row=row, column=3).value = temp_time  # user_info[2]時間 excelの書式の都合

    workbook.save(thread_list_filepath)

    # 提出しているかどうかをスレッドごとに確認してafter_check_filepath(excel)に書き込む
    submit_info_write_excel(
        thread_title_original,
        user_info[0],
        after_check_filepath,
        after_check_file_sheet,
        user_info[2],
        submit_deadline,
        delay_deadline,
    )


# 提出していない人に0を書き込む関数
def write_zero_to_excel(after_check_filepath, after_check_file_sheet, title):
    wb = openpyxl.load_workbook(after_check_filepath)

    if after_check_file_sheet not in wb.sheetnames:
        print("シートが存在しない")
        return

    sheet = wb[after_check_file_sheet]

    # 課題名に一致する列を探す
    for i in range(9, 25):  # 課題（横）のループ
        if (
            (str(sheet.cell(row=2, column=i).value) in str(title))
            or (str(title) in (str(sheet.cell(row=2, column=i).value)))
            or (str(title) == (str(sheet.cell(row=2, column=i).value)))
        ):
            print(str(sheet.cell(row=2, column=i).value))
            # 学生の名前に一致する行を探す
            for j in range(3, 98):  # 学生（縦）のループ
                if not sheet.cell(row=j, column=i).value:
                    sheet.cell(row=j, column=i).value = 0
            break
    wb.save(after_check_filepath)


def main():
    driver = webdriver.Chrome()

    student_id = os.getenv("STUDENT_ID")
    password = os.getenv("PASSWORD")

    login_to_moodle(driver, student_id, password)

    # 先生に提出するための確認済みのexcelファイル
    # TODO: もらった名簿をもとにファイルを作成する
    after_check_file_name = "DB2023課題チェック名簿1201"
    after_check_filepath = "./excel/" + after_check_file_name.strip() + ".xlsx"

    # 提出するexcelファイルの各回のシート名
    # TODO: 講義毎に変更する
    after_check_file_sheet = "DB2024第五章受講者リスト0212"
    course_name = "データベース 2023 KA126001"
    section_title = "第5章課題"

    # TODO:　スレッド毎に変更する
    title = "課題５－１　Task5-1"
    # 提出締め切り
    submit_deadline = "2024-02-11 23:59:59"
    # 遅延締め切り
    delay_deadline = "2024-02-11 23:59:59"

    access_assignment(driver, course_name, section_title, title)

    sec_title_for_filepath = section_title.translate(
        str.maketrans(
            {
                "?": "",
                "？": "",
                "/": "",
                "\\": "",
                ":": "",
                "*": "",
                "<": "",
                ">": "",
                "|": "",
            }
        )
    )
    thread_list_filepath = "./excel/" + sec_title_for_filepath.strip() + ".xlsx"

    wb = create_or_load_excel(thread_list_filepath)
    elements_with_class = threads_list(driver, "forum-post-container mb-2")
    thread_count = len(elements_with_class)

    for i in range(thread_count):
        sleep(1)
        process_thread(
            elements_with_class[i],
            wb,
            thread_list_filepath,
            after_check_filepath,
            after_check_file_sheet,
            submit_deadline,
            delay_deadline,
        )
    print("遅延提出の確認まで完了しました。")
    write_zero_to_excel(after_check_filepath, after_check_file_sheet, title)

    driver.close()


if __name__ == "__main__":
    main()
